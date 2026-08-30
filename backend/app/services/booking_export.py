from collections import Counter
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings
from app.models.booking import Booking, BookingStatus

STATUS_LABELS = {
    BookingStatus.WAITING_PAYMENT: "Ожидает оплаты",
    BookingStatus.CONFIRMED: "Подтверждена",
    BookingStatus.CANCELLED: "Отменена",
    BookingStatus.COMPLETED: "Выполнена",
    BookingStatus.NO_SHOW: "Не приехал",
}
HEADER_FILL = PatternFill("solid", fgColor="D71920")
SUBHEADER_FILL = PatternFill("solid", fgColor="181818")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _status_label(status: BookingStatus) -> str:
    return STATUS_LABELS.get(status, status.value)


def _local_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(ZoneInfo(settings.business_timezone)).replace(tzinfo=None)


def _style_header(worksheet: Worksheet, row: int, start: int, end: int) -> None:
    for column in range(start, end + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def build_booking_workbook(bookings: list[Booking]) -> bytes:
    workbook = Workbook()
    details = workbook.active
    if details is None:
        raise RuntimeError("Workbook has no active worksheet")
    details.title = "Заказы"
    details.freeze_panes = "A2"
    details.sheet_view.showGridLines = False

    headers = [
        "ID",
        "Создана",
        "Дата визита",
        "Начало",
        "Окончание",
        "Клиент",
        "Телефон",
        "Автомобиль",
        "Цвет",
        "Услуга",
        "Статус",
    ]
    details.append(headers)
    _style_header(details, 1, 1, len(headers))
    details.row_dimensions[1].height = 24

    for booking in bookings:
        details.append(
            [
                str(booking.id),
                _local_datetime(booking.created_at),
                booking.date,
                booking.start_time,
                booking.end_time,
                booking.client_name,
                booking.client_phone,
                booking.vehicle_model,
                booking.vehicle_color,
                booking.service_name,
                _status_label(booking.status),
            ]
        )

    if bookings:
        table = Table(displayName="BookingsTable", ref=f"A1:K{len(bookings) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        details.add_table(table)
    details.auto_filter.ref = f"A1:K{max(1, len(bookings) + 1)}"
    details.column_dimensions["A"].width = 38
    details.column_dimensions["B"].width = 20
    details.column_dimensions["C"].width = 15
    details.column_dimensions["D"].width = 12
    details.column_dimensions["E"].width = 12
    details.column_dimensions["F"].width = 24
    details.column_dimensions["G"].width = 20
    details.column_dimensions["H"].width = 26
    details.column_dimensions["I"].width = 16
    details.column_dimensions["J"].width = 28
    details.column_dimensions["K"].width = 18
    for row in range(2, len(bookings) + 2):
        details.cell(row=row, column=2).number_format = "dd.mm.yyyy hh:mm"
        details.cell(row=row, column=3).number_format = "dd.mm.yyyy"
        details.cell(row=row, column=4).number_format = "hh:mm"
        details.cell(row=row, column=5).number_format = "hh:mm"

    statistics = workbook.create_sheet("Статистика")
    statistics.sheet_view.showGridLines = False
    statistics.merge_cells("A1:E1")
    statistics["A1"] = "Статистика заказов NazarovGroup"
    statistics["A1"].fill = HEADER_FILL
    statistics["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    statistics["A1"].alignment = Alignment(vertical="center")
    statistics.row_dimensions[1].height = 30
    statistics["A3"] = "Всего заказов"
    statistics["B3"] = len(bookings)
    statistics["A3"].font = Font(bold=True)

    statistics["A5"] = "Статус"
    statistics["B5"] = "Количество"
    statistics["D5"] = "Услуга"
    statistics["E5"] = "Количество"
    for cell in (statistics["A5"], statistics["B5"], statistics["D5"], statistics["E5"]):
        cell.fill = SUBHEADER_FILL
        cell.font = HEADER_FONT

    status_counts = Counter(booking.status for booking in bookings)
    for row, status in enumerate(BookingStatus, start=6):
        statistics.cell(row=row, column=1, value=_status_label(status))
        statistics.cell(row=row, column=2, value=status_counts.get(status, 0))

    service_counts = Counter(booking.service_name for booking in bookings)
    for row, (service_name, count) in enumerate(service_counts.most_common(), start=6):
        statistics.cell(row=row, column=4, value=service_name)
        statistics.cell(row=row, column=5, value=count)

    last_row = max(10, len(service_counts) + 5)
    for row in statistics.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    for column, width in {"A": 22, "B": 14, "C": 4, "D": 32, "E": 14}.items():
        statistics.column_dimensions[column].width = width

    workbook.properties.title = "Статистика заказов NazarovGroup"
    workbook.properties.creator = "NazarovGroup"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
